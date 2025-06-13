from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generar_excel(productos):
    wb = Workbook()  # Crear un nuevo libro de trabajo
    ws = wb.active  # Seleccionar la hoja activa
    ws.title = 'Productos'  # Establecer el título de la hoja

    # Definir estilos
    header_fill = PatternFill(start_color='A9A9F5', end_color='A9A9F5', fill_type='solid')  # Azul claro
    first_col_fill = PatternFill(start_color='C38EC7', end_color='C38EC7', fill_type='solid')  # Morado
    alt_row_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Gris claro
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Encabezados
    headers = ['ID', 'Nombre', 'Serial', 'Cantidad']
    ws.append(headers)
    for col, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.border = border

    # Datos
    row_num = 2
    for producto in productos:
        for equipo in producto.equipo_set.all():
            ws.append([producto.id, producto.nombre, equipo.serial, producto.equipo_set.count()])
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                # Primera columna morada
                if col_num == 1:
                    cell.fill = first_col_fill
                # Filas alternas gris claro
                elif row_num % 2 == 0:
                    cell.fill = alt_row_fill
                cell.border = border
            row_num += 1

    return wb  # Retornar el libro de trabajo para su uso posterior